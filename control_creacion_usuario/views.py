from email.mime.image import MIMEImage
import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,update_session_auth_hash, login as auth_login, logout as auth_logout
from django.shortcuts import redirect
from django.contrib.auth.models import User
from django.http import FileResponse, HttpResponse,JsonResponse
from core.models import UserActivity
from formulario.models import ProtocoloSolicitud,Registro_designio
from django.utils import timezone
from openpyxl import Workbook
# Create your views here.
from django.views.decorators.csrf import csrf_exempt
from .forms import ImagenForm,PDFForm
from .models import Imagen_sig,PDF_sig
from django.shortcuts import get_object_or_404, redirect
from django.core.paginator import Paginator, Page
from datetime import datetime, timedelta
from django.utils.timezone import make_aware
from formulario.models import *
from django.db.models import Q,F, Sum,Count
import os
from django.contrib import messages
from django.utils.timezone import now
import smtplib
from smtplib import *
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import pytz
from email.utils import encode_rfc2231
from django.core.files.storage import default_storage
from django.utils.html import escape
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from .pdf_generator import *
from tareas.models import *
from django.db import transaction


ESTADO = [
    ('RECIBIDO', 'RECIBIDO'),
    ('EN PROCESO', 'EN PROCESO'),
    ('EJECUTADO', 'EJECUTADO'),
    ('RECHAZADO', 'RECHAZADO')
]

LIMITE_DE_DIA = [
    ('', ''),
    ('L', '1 - 2 Días Hábiles'),
    ('M', '2 - 4 Días Hábiles'),
    ('A', '3 - 5 Días Hábiles'),
    ('P', 'Especificar días manualmente'),
]

OPCIONES = {
    'ESTADO': ESTADO,
    'LIMITE_DE_DIA': LIMITE_DE_DIA
}


@csrf_exempt
def login(request):
    if request.user.is_authenticated:
        return redirect('control')

    if request.method == 'POST':
        email = request.POST.get('email')
        email = email.lower()
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)

        if user is not None:
            auth_login(request, user)
            return redirect('solicitude_llegadas')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')

    return render(request, 'Login.html')

def download_excel(request):
    # Crear un nuevo libro de trabajo de Excel y agregar datos
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'PAGINA'
    sheet['B1'] = 'DEPARTAMENTO'
    sheet['C1'] = 'MES-DIA-AÑO-HORA'

    sheet.column_dimensions['A'].width = 30  # Ancho de la columna B
    sheet.column_dimensions['B'].width = 50  # Ancho de la columna C
    sheet.column_dimensions['C'].width = 30  # Ancho de la columna C


    row = 2  # Fila inicial para los datos
    activities = UserActivity.objects.all()

    for activity in activities:
        sheet.cell(row=row, column=1).value = activity.page
        sheet.cell(row=row, column=2).value = activity.departamento
        sheet.cell(row=row, column=3).value = activity.timestamp.strftime('%m-%d-%Y-%H:%M')
        row += 1  # Incrementar la fila para el próximo registro
    # Configurar la respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Historial_de_visitas.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    workbook.save(response)

    return response

def Historial_Visitas(request):
    Historial = UserActivity.objects.all()
    data = {
    'historial': Historial,
    }
    return render(request,'Historial_Visitas.html',data)
from django.http import JsonResponse
from django.db.models.functions import ExtractYear
from django.utils.timezone import now
@login_required(login_url='/login/')

def solicitude_llegadas(request):
    usuarios = User.objects.all()
    filtro_anio = request.GET.get('anio')
    filtro_id = request.GET.get('id', '').strip()

    # Obtener solicitudes base según usuario
    if request.user.is_superuser:
        solicitudes = ProtocoloSolicitud.objects.all()
    else:
        solicitudes = ProtocoloSolicitud.objects.filter(profesional=request.user)

    # Filtrar por año si existe filtro
    if filtro_anio:
        solicitudes = solicitudes.filter(fecha_L__year=filtro_anio)

    # Filtrar por ID si existe filtro
    if filtro_id:
        solicitudes = solicitudes.filter(id__icontains=filtro_id)

    # Generar datos de solicitudes con info extra
    solicitudes_data = []
    for solicitud in solicitudes:
        numero_designios = Registro_designio.objects.filter(protocolo=solicitud).count()

        if solicitud.fecha_T:
            dias_restantes = "Trabajo terminado"
        elif solicitud.fecha_L:
            total_segundos = (solicitud.fecha_L - now()).total_seconds()

            if total_segundos < 0:
                total_segundos = abs(total_segundos)
                dias_pasados = int(total_segundos // (24 * 3600))
                horas_pasadas = int((total_segundos % (24 * 3600)) // 3600)
                minutos_pasados = int((total_segundos % 3600) // 60)
                dias_restantes = f"Pasada por {dias_pasados} días, {horas_pasadas} horas y {minutos_pasados} minutos"
            else:
                dias_habiles_restantes = calcular_dias_habiles(now().date(), solicitud.fecha_L.date())
                horas_restantes = int((total_segundos % (24 * 3600)) // 3600)
                minutos_restantes = int((total_segundos % 3600) // 60)
                if dias_habiles_restantes > 1:
                    dias_restantes = f"Te quedan {dias_habiles_restantes-1} días hábiles"
                elif horas_restantes > 0:
                    dias_restantes = f"Te quedan {horas_restantes} horas y {minutos_restantes} minutos"
                else:
                    dias_restantes = f"Te quedan {minutos_restantes} minutos"
        else:
            dias_restantes = "Sin fecha límite"

        # Manejo seguro del campo "solicitante"
        solicitante_obj = getattr(solicitud, 'solicitante', None)
        solicitante_nombre = solicitante_obj.username if solicitante_obj else ''

        solicitudes_data.append({
            'id': solicitud.id,
            'orden_trabajo': getattr(solicitud, 'orden_trabajo', ''),
            'solicitante': solicitante_nombre,
            'fecha_llegada': solicitud.fecha_L.strftime('%Y-%m-%d') if solicitud.fecha_L else '',
            'departamento': getattr(solicitud, 'departamento', ''),
            'numero_designios': numero_designios,
            'dias_restantes': dias_restantes,
            # Puedes seguir agregando más campos aquí
        })

    # Extraer años para el filtro dinámico
    if request.user.is_superuser:
        años = ProtocoloSolicitud.objects.annotate(year=ExtractYear('fecha_L')).values_list('year', flat=True).distinct()
    else:
        años = ProtocoloSolicitud.objects.filter(profesional=request.user).annotate(year=ExtractYear('fecha_L')).values_list('year', flat=True).distinct()

    años = sorted([a for a in años if a is not None], reverse=True)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'solicitudes': solicitudes_data})

    return render(request, 'solicitude_llegadas.html', {'anios': años})


def calcular_dias_habiles(fecha_inicio, fecha_fin):
    dias_habiles = 0
    dia_actual = fecha_inicio

    while dia_actual <= fecha_fin:
        # Si el día actual no es sábado (5) ni domingo (6), lo contamos
        if dia_actual.weekday() < 5:  # 0 = Lunes, ..., 4 = Viernes
            dias_habiles += 1
        dia_actual += timedelta(days=1)
    
    return dias_habiles

# Relacionar Funcionario con User basado en nombres y apellidos
def obtener_usuario_por_funcionario(funcionario):
    return User.objects.filter(
        first_name=funcionario.nombre, last_name=funcionario.apellido
    ).first()

@login_required(login_url='/login/')
def control(request):
    # Obtener todas las solicitudes y tareas
    total_solicitudes = ProtocoloSolicitud.objects.count()
    total_tareas = Tarea.objects.filter(completada=True).count()

    # Contar solicitudes por estado
    estados = ProtocoloSolicitud.objects.values("estado").annotate(total=Count("estado"))
    estado_counts = {estado["estado"]: estado["total"] for estado in estados}

    estados_posibles = ["RECIBIDO", "EN PROCESO", "EJECUTADO", "RECHAZADO"]
    for estado in estados_posibles:
        estado_counts.setdefault(estado, 0)

    profesionales_solicitudes = (
        ProtocoloSolicitud.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name", "profesional_id")
        .annotate(
            trabajo_propio=Sum("valor_de_trabajo_funcionario"),
            total_solicitudes=Count("id")
        )
    )

    profesionales_apoyos = (
        Apoyo_Protocolo.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name", "profesional_id")
        .annotate(
            trabajo_apoyo=Sum("valor_de_trabajo"),
            total_solicitudes_apoyo=Count("id")
        )
    )

    tareas_completadas = (
        Tarea.objects.filter(completada=True)
        .values("funcionario__nombre", "funcionario__apellido", "funcionario_id")
        .annotate(total_tareas=Count("id"))
    )

    tareas_con_apoyo = (
        Tarea.objects.filter(completada=True, apoyo__isnull=False)
        .values("funcionario_id")
        .annotate(total_apoyos=Count("id"))
    )

    puntajes_por_profesional = {}

    for item in profesionales_solicitudes:
        full_name = f"{item['profesional__first_name']} {item['profesional__last_name']}"
        puntajes_por_profesional[item["profesional_id"]] = {
            "name": full_name,
            "trabajo_propio": float(item["trabajo_propio"] or 0),
            "trabajo_apoyo": 0,
            "total_solicitudes": item["total_solicitudes"],
            "total_solicitudes_apoyo": 0,
            "total_tareas": 0
        }
    
    solicitudes_per_profesional = (
        ProtocoloSolicitud.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    labels = [
        f"{item['profesional__first_name']} {item['profesional__last_name']}"
        for item in solicitudes_per_profesional
    ]

    for item in profesionales_apoyos:
        full_name = f"{item['profesional__first_name']} {item['profesional__last_name']}"
        if item["profesional_id"] in puntajes_por_profesional:
            puntajes_por_profesional[item["profesional_id"]]["trabajo_apoyo"] = float(item["trabajo_apoyo"] or 0)
            puntajes_por_profesional[item["profesional_id"]]["total_solicitudes_apoyo"] = item["total_solicitudes_apoyo"]
        else:
            puntajes_por_profesional[item["profesional_id"]] = {
                "name": full_name,
                "trabajo_propio": 0,
                "trabajo_apoyo": float(item["trabajo_apoyo"] or 0),
                "total_solicitudes": 0,
                "total_solicitudes_apoyo": item["total_solicitudes_apoyo"],
                "total_tareas": 0
            }

    for item in tareas_completadas:
        funcionario = Funcionario.objects.get(id=item["funcionario_id"])
        user = obtener_usuario_por_funcionario(funcionario)
        if user:
            if user.id in puntajes_por_profesional:
                puntajes_por_profesional[user.id]["total_tareas"] = item["total_tareas"]
            else:
                full_name = f"{funcionario.nombre} {funcionario.apellido}"
                puntajes_por_profesional[user.id] = {
                    "name": full_name,
                    "trabajo_propio": 0,
                    "trabajo_apoyo": 0,
                    "total_solicitudes": 0,
                    "total_solicitudes_apoyo": 0,
                    "total_tareas": item["total_tareas"]
                }

    for item in tareas_con_apoyo:
        funcionario = Funcionario.objects.get(id=item["funcionario_id"])
        user = obtener_usuario_por_funcionario(funcionario)
        if user:
            if user.id in puntajes_por_profesional:
                puntajes_por_profesional[user.id]["total_apoyos_tareas"] = item["total_apoyos"]
            else:
                full_name = f"{funcionario.nombre} {funcionario.apellido}"
                puntajes_por_profesional[user.id] = {
                    "name": full_name,
                    "trabajo_propio": 0,
                    "trabajo_apoyo": 0,
                    "total_solicitudes": 0,
                    "total_solicitudes_apoyo": 0,
                    "total_tareas": 0,
                    "total_apoyos_tareas": item["total_apoyos"]
                }

    sorted_puntajes = sorted(
        puntajes_por_profesional.values(),
        key=lambda x: x["trabajo_propio"] + x["trabajo_apoyo"] + x["total_tareas"] + x.get("total_apoyos_tareas", 0),
        reverse=True
    )

    labels = [item["name"] for item in sorted_puntajes]
    trabajo_propio_data = [item["trabajo_propio"] for item in sorted_puntajes]
    trabajo_apoyo_data = [item["trabajo_apoyo"] for item in sorted_puntajes]
    total_solicitudes_data = [item["total_solicitudes"] for item in sorted_puntajes]
    total_solicitudes_apoyo_data = [item["total_solicitudes_apoyo"] for item in sorted_puntajes]
    total_tareas_data = [item["total_tareas"] for item in sorted_puntajes]
    total_apoyos_tareas_data = [item.get("total_apoyos_tareas", 0) for item in sorted_puntajes]

    labels_json = json.dumps(labels)
    total_unitario_solicitudes_json = json.dumps(total_solicitudes_data)
    total_unitario_solicitudes_apoyo_json = json.dumps(total_solicitudes_apoyo_data)
    total_tareas_json = json.dumps(total_tareas_data)
    total_apoyos_tareas_json = json.dumps(total_apoyos_tareas_data)
    trabajo_porcentual_propio_json = json.dumps(trabajo_propio_data)
    trabajo_porcentual_apoyo_json = json.dumps(trabajo_apoyo_data)

    context = {
        "total_solicitudes": total_solicitudes,
        "total_tareas": total_tareas,
        "labels_json": labels_json,
        "total_unitario_solicitudes_json": total_unitario_solicitudes_json,
        "total_unitario_solicitudes_apoyo_json": total_unitario_solicitudes_apoyo_json,
        "total_tareas_json": total_tareas_json,
        "total_apoyos_tareas_json": total_apoyos_tareas_json,
        "trabajo_porcentual_propio_json": trabajo_porcentual_propio_json,
        "trabajo_porcentual_apoyo_json": trabajo_porcentual_apoyo_json,
        "en_proceso": estado_counts.get("EN PROCESO", 0),
        "ejecutado": estado_counts.get("EJECUTADO", 0),
        "rechazado": estado_counts.get("RECHAZADO", 0),
    }

    return render(request, "Control.html", context)


def cambiar_contraseña(request):
    if request.method == 'POST':
            password1 = request.POST['password1']
            password2 = request.POST['password2']

            if password1 == password2:
                user = User.objects.get(id=request.user.id)
                user.set_password(password1)
                user.save()

                # Mantener al usuario autenticado después del cambio de contraseña
                update_session_auth_hash(request, user)

                return redirect('control')
            else:
                messages.error(request, 'Las contraseñas no coinciden. Inténtalo de nuevo.')

    return render(request, 'cambiar_contraseña.html')

def logout(request):
    auth_logout(request)
    return redirect('control')

from django.core.paginator import Paginator
from django.urls import reverse
def Gestion_imagen(request):
    usuarios = User.objects.all().order_by('username')  # Para el filtro

    usuario_id = request.GET.get('usuario')  # Puede ser None, "", o id string

    imagenes = Imagen_sig.objects.all().order_by('-id')
    usuario_filtrado = None
    nombre_usuario = "Todos"

    # Caso 1: No viene filtro -> filtro por usuario logueado por defecto
    if usuario_id is None:
        usuario_id = str(request.user.id)
        try:
            usuario_filtrado = int(usuario_id)
            imagenes = imagenes.filter(usuario_id=usuario_filtrado)
            usuario_obj = User.objects.get(id=usuario_filtrado)
            nombre_usuario = usuario_obj.username
        except (User.DoesNotExist, ValueError):
            usuario_filtrado = None
            nombre_usuario = "Desconocido"

    # Caso 2: Viene filtro vacío "" -> mostrar todas las imágenes
    elif usuario_id == '':
        usuario_filtrado = None
        nombre_usuario = "Todos"
        # imagenes ya tiene todas las imágenes, no filtramos

    # Caso 3: Viene id usuario -> filtrar por ese usuario
    else:
        try:
            usuario_filtrado = int(usuario_id)
            imagenes = imagenes.filter(usuario_id=usuario_filtrado)
            usuario_obj = User.objects.get(id=usuario_filtrado)
            nombre_usuario = usuario_obj.username
        except (User.DoesNotExist, ValueError):
            usuario_filtrado = None
            nombre_usuario = "Desconocido"
            # mostrar todas las imágenes
            imagenes = Imagen_sig.objects.all().order_by('-id')

    paginator = Paginator(imagenes, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = ImagenForm(request.POST, request.FILES)
        if form.is_valid():
            imagen = form.save(commit=False)
            imagen.usuario = request.user
            imagen.save()
            # redirigir manteniendo filtro actual
            redirect_url = reverse('Gestion_imagen')  # ✅ convierte el nombre en una URL real
            query_params = f'?usuario={usuario_id}' if usuario_id is not None else ''
            return redirect(f"{redirect_url}{query_params}")

    else:
        form = ImagenForm()

    context = {
        'usuarios': usuarios,
        'usuario_filtrado': str(usuario_filtrado) if usuario_filtrado is not None else '',
        'nombre_usuario': nombre_usuario,
        'page_obj': page_obj,
        'form': form,
    }
    return render(request, 'Gestion_imagen.html', context)



@csrf_exempt
def actualizar_estado_solicitud(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        nuevo_estado = request.POST.get('estado')

        try:
            solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)
            if nuevo_estado == "EJECUTADO" and solicitud.profesional:
                solicitud.estado = nuevo_estado
                solicitud.fecha_T = timezone.now()
                solicitud.save()
            elif nuevo_estado == "RECHAZADO":
                solicitud.estado = nuevo_estado
                solicitud.save()
            else:
                solicitud.estado = nuevo_estado
                solicitud.save()

            # 🔥 Notificar a WebSockets 🔥
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "solicitudes",
                {"type": "send_update", "message": "actualizar"}
            )

            return JsonResponse({'success': True, 'message': 'Estado actualizado correctamente'})

        except ProtocoloSolicitud.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Solicitud no encontrada'})

    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

@csrf_exempt
def actualizar_profesional(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        nuevo_profesional_id = request.POST.get('profesional')
        motivo = request.POST.get('motivo')
        solicitud = get_object_or_404(ProtocoloSolicitud, id=solicitud_id)

        



        if not nuevo_profesional_id or not solicitud_id:
            return JsonResponse({'success': False, 'message': 'Necesita'})

        if solicitud.profesional:
            Reg = Registro_designio(
                objetivos = motivo,
                protocolo = solicitud,
                profesional = solicitud.profesional,
                
                )
            Reg.save()
            
        nuevo_profesional = get_object_or_404(User, id=nuevo_profesional_id)
        solicitud.profesional = nuevo_profesional
        solicitud.fecha_D = timezone.now()
        solicitud.save()

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
                "solicitudes",
                {"type": "send_update", "message": "actualizar"}
            )
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido'})

def calcular_fecha_limite(inicio, dias):
    fecha_limite = inicio
    dias_restantes = dias

    while dias_restantes > 0:
        fecha_limite += timedelta(days=1)
        if fecha_limite.weekday() < 5:  # Lunes=0, Domingo=6
            dias_restantes -= 1
    
    return fecha_limite

@csrf_exempt
def actualizar_limite(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        tipo_limite = request.POST.get('nuevoLimite')
        fecha_completa = request.POST.get('fecha')  # Fecha completa enviada desde el frontend

        try:
            # Validar si la solicitud existe
            solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)

            if tipo_limite == 'P':
                if not fecha_completa:
                    return JsonResponse({'success': False, 'message': 'Debe proporcionar una fecha válida para el tipo P'})
                try:
                    # Convertir la fecha completa ISO 8601 a datetime
                    fecha_limite = datetime.strptime(fecha_completa, "%Y-%m-%dT%H:%M:%S.%fZ")
                    # Asegurarse de que sea consciente de la zona horaria UTC y convertir a la zona horaria local
                    fecha_limite = make_aware(fecha_limite, timezone=pytz.UTC)
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'Formato de fecha inválido'})
            else:
                # Calcular los días límite según el tipo
                if tipo_limite == 'A':
                    dias_limite = 5
                elif tipo_limite == 'M':
                    dias_limite = 4
                elif tipo_limite == 'L':
                    dias_limite = 1
                else:
                    return JsonResponse({'success': False, 'message': 'Tipo de límite no reconocido'})

                # Calcular la fecha límite basada en días
                fecha_limite = calcular_fecha_limite(timezone.now(), dias_limite)

            # Actualizar la solicitud con el nuevo tipo de límite y fecha límite
            solicitud.tipo_limite = tipo_limite
            solicitud.fecha_L = fecha_limite
            solicitud.save()
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "solicitudes",
                {"type": "send_update", "message": "actualizar"}
            )

            return JsonResponse({'success': True, 'message': 'Límite actualizado correctamente'})

        except ProtocoloSolicitud.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'La solicitud no existe'})

    return JsonResponse({'success': False, 'message': 'Método no permitido'})

def eliminar_imagen(request, imagen_id):
    imagen = get_object_or_404(Imagen_sig, pk=imagen_id)
    
    if request.method == 'POST':
        # Guarda la ruta del archivo de la imagen
        ruta_archivo = imagen.archivo_adjunto.path
        # Elimina la imagen de la base de datos
        imagen.delete()
        # Elimina el archivo de la imagen del sistema de archivos
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
        return redirect('Gestion_imagen')  # Redirige a la vista de gestión de imágenes después de eliminar la imagen
    
    return redirect('Gestion_imagen')  # Redirige a la vista de gestión de imágenes si la solicitud no es de tipo POST

def Gestion_pdf(request):
    if request.method == 'POST':
        form = PDFForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('Gestion_pdf')  # Redirige a la misma vista después de guardar la imagen
    else:
        form = PDFForm()

    # Obtener todas las imágenes
    pdf = PDF_sig.objects.all()
    paginator = Paginator(pdf, 6)  # Muestra 6 imágenes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Gestion_PDF.html', {'form': form, 'PDF': pdf,'page_obj': page_obj})

def Calculor_de_trabajo():

    total_ejecutado = ProtocoloSolicitud.objects.filter(
    estado="EJECUTADO",
    fecha_D__isnull=False).count()

    print("Número total de solicitudes válidas:", total_ejecutado)


    total_ejecutado_a_tiempo = ProtocoloSolicitud.objects.filter(
        estado="EJECUTADO",
        fecha_D__isnull=False,
        fecha_T__isnull=False,                                        
        fecha_T__lte=F('fecha_L')  # fecha_T menor o igual que fecha_L
    ).count()

    print("Número total de solicitudes válidas a tiempo:", total_ejecutado_a_tiempo)


    solicitudes = ProtocoloSolicitud.objects.filter(
        estado="EJECUTADO",
        fecha_D__isnull=False,
        fecha_T__isnull=False
    )

    # Calcular el total de segundos
    total_dias = 0
    for s in solicitudes:
        diferencia = s.fecha_D  -  s.fecha_T # Restar los DateTimeField, obteniendo un timedelta
        total_dias += round(diferencia.total_seconds() / 86400, 3)  # Convertir segundos a días y limitar a 3 decimales

    total_dias = round(total_dias,3)
    print("Total de días entre fecha_D y fecha_T:", total_dias)


            
    if total_ejecutado > 0:
        tpr = total_ejecutado /  total_dias
        tpr = round(tpr,3)

    else:
        tpr = 0  # Evitar división por cero si no hay solicitudes ejecutadas

    print("Tiempo Promedio de Resolución (TPR) en Dias:", tpr)

    if total_ejecutado_a_tiempo > 0:
        tdc = (total_ejecutado /   total_ejecutado_a_tiempo )*100
        tdc = round(tdc,3)

    else:
        tdc = 0

    print("Tasa de cumplimiento:", tdc)

    return tdc,tpr

@csrf_exempt  # Solo usar esto si estás probando; para producción, configura CSRF correctamente
def Envio_de_correo(request):
    if request.method == 'POST':
        user = request.user
        email_input = request.POST.get('email', '').strip()

        # Convertir la entrada en una lista separando por comas, eliminando espacios y vacíos
        emails = [email.strip() for email in email_input.split(',') if email.strip()] if email_input else []

        archivos = request.FILES.getlist('files')
        total_size = sum(archivo.size for archivo in archivos)

        urls_archivos = []
        archivos_adjuntos = []
        dominio = f"{request.scheme}://{request.get_host()}"


        if user.is_superuser:
            try:
                # Obtener los datos
                message = request.POST.get('message', '')
                formatted_message = escape(message).replace("\n", "<br>")
                ficha_id = request.POST.get('ficha_id')
                Protocolo = ProtocoloSolicitud.objects.get(id=ficha_id)
                Protocolo.enviado_correo = True
                profesional = Protocolo.profesional
                Protocolo.estado = 'EN PROCESO'
                Protocolo.save()

                # Verificar el tamaño de los archivos
                if total_size > 10 * 1024 * 1024:  # Más de 10 MB
                    for archivo in archivos:
                        archivo_link = Archivo_Link.objects.create(protocolo=Protocolo, archivo=archivo)
                        url_relativa = default_storage.url(archivo_link.archivo.name)
                        urls_archivos.append(f"{dominio}{url_relativa}")

                else:  # Menor o igual a 10 MB
                    for archivo in archivos:
                        archivo_adjunto = MIMEApplication(archivo.read())
                        nombre_archivo = encode_rfc2231(archivo.name, 'utf-8')
                        archivo_adjunto.add_header(
                            'Content-Disposition',
                            'attachment',
                            filename=nombre_archivo
                        )
                        archivos_adjuntos.append(archivo_adjunto)

                # Generar PDF

                data = {
                    'Protocolo': Protocolo,
                }

                nombre_ficha = "Solicitud" + str(Protocolo.id) + "_" + str(Protocolo)

                # Generar el PDF
                pdf_path, status = save_pdf_3(data, nombre_ficha)

                # Leer el archivo generado en modo binario
                with open(pdf_path, "rb") as pdf_file:
                    archivo_pdf = pdf_file.read()

                # Configuración del correo
                superusers = User.objects.filter(is_superuser=True).exclude(username="osvaldo.moya").values_list('email', flat=True)
                superuser_emails = list(superusers)

                mi_correo = f'{user.username}@munivalpo.cl'.strip()
                asunto = f'Solicitud N°{Protocolo.codigo} Asignada'
                mensaje = MIMEMultipart()
                mensaje['From'] = mi_correo
                destinatarios = list(set([profesional.email] + emails))
                mensaje['To'] = ', '.join(destinatarios)
                mensaje['Subject'] = asunto

                bcc_destinatarios = [mi_correo]

                # Cargar la firma
                firma_path = os.path.join('media/assets/Firma', f'{user.username}.png')
                if os.path.exists(firma_path):
                    with open(firma_path, 'rb') as firma_file:
                        firma_img = MIMEImage(firma_file.read())
                        firma_img.add_header('Content-ID', '<firma>')
                        mensaje.attach(firma_img)

                # Generar el contenido HTML
                html_archivos = ""
                if urls_archivos:
                    html_archivos = f"""
                    <p>Los siguientes archivos superan el límite de 10 MB y están disponibles en los siguientes enlaces:</p>
                    <ul>
                        {''.join(f'<li><a href="{url}" target="_blank">{url}</a></li>' for url in urls_archivos)}
                    </ul>
                    """
                # else:
                #     html_archivos = "<p>Los archivos están adjuntos al correo.</p>"

                html_content = f"""
                <html>
                    <body>
                        <p>{formatted_message}</p>
                        <br>
                        {html_archivos}
                        <br>
                        <img src="cid:firma" alt="Firma" width="600" height="auto" />
                    </body>
                </html>
                """
                mensaje.attach(MIMEText(html_content, 'html'))

                # Adjuntar PDF
                pdf_adjunto = MIMEApplication(archivo_pdf)
                pdf_adjunto.add_header('Content-Disposition', 'attachment', filename='Ficha_de_protocolo.pdf')
                mensaje.attach(pdf_adjunto)

                # Adjuntar archivos menores a 10 MB
                for archivo_adjunto in archivos_adjuntos:
                    mensaje.attach(archivo_adjunto)

                # Configuración del servidor SMTP
                smtp_server = 'mail.munivalpo.cl'
                smtp_port = 587
                smtp_usuario = f'servervalpo\\{user.username}'
                smtp_contrasena = encotra_contraseña(user.username)

                # Enviar el correo
                server = smtplib.SMTP(smtp_server, smtp_port)
                server.starttls()
                server.login(smtp_usuario, smtp_contrasena)
                server.sendmail(
                    mi_correo,
                    destinatarios + bcc_destinatarios,  # Incluir destinatarios normales y BCC
                    mensaje.as_string()
                )

                
                server.quit()
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                        "solicitudes",
                        {"type": "send_update", "message": "actualizar"}
                    )
                

                return JsonResponse({'success': True})
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()  # Obtiene el error detallado
                return JsonResponse({'success': False, 'error': str(e), 'traceback': error_trace}, status=555)


        else:
            try:
                message = request.POST.get('message', '')
                formatted_message = escape(message).replace("\n", "<br>")
                ficha_id = request.POST.get('ficha_id')
                Protocolo = ProtocoloSolicitud.objects.get(id=ficha_id)
                Protocolo.enviado_correo_t = True
                Protocolo.fecha_T = timezone.now()
                profesional = Protocolo.profesional
                Protocolo.estado = 'EJECUTADO'
                solicitante = Protocolo.corre_solicitante

                Protocolo.save()

                # Verificar el tamaño de los archivos
                if total_size > 10 * 1024 * 1024:  # Más de 10 MB
                    for archivo in archivos:
                        archivo_link = Archivo_Link.objects.create(protocolo=Protocolo, archivo=archivo)
                        url_relativa = default_storage.url(archivo_link.archivo.name)
                        urls_archivos.append(f"{dominio}{url_relativa}")

                else:  # Menor o igual a 10 MB
                    for archivo in archivos:
                        archivo_adjunto = MIMEApplication(archivo.read())
                        nombre_archivo = encode_rfc2231(archivo.name, 'utf-8')
                        archivo_adjunto.add_header(
                            'Content-Disposition',
                            'attachment',
                            filename=nombre_archivo
                        )
                        archivos_adjuntos.append(archivo_adjunto)

                # Generar PDF
                data = {
                    'Protocolo': Protocolo,
                }

                nombre_ficha = "Solicitud" + str(Protocolo.id) + "_" + str(Protocolo)

                # # Generar el PDF
                pdf_path, status = save_pdf_3(data, nombre_ficha)

                # # Leer el archivo generado en modo binario
                with open(pdf_path, "rb") as pdf_file:
                    archivo_pdf = pdf_file.read()


                if not status:
                    print("----------------")
                    print("Error al generar PDF")
                    print("----------------")
                    return HttpResponse("Error al generar PDF")

                # Configuración del correo
                superusers = User.objects.filter(is_superuser=True).exclude(username="osvaldo.moya").values_list('email', flat=True)
                superuser_emails = list(superusers)

                mi_correo = f'{user.username}@munivalpo.cl'.strip()
                asunto = f'Solicitud N°{Protocolo.codigo} Asignada'
                mensaje = MIMEMultipart()
                mensaje['From'] = mi_correo
                destinatarios = list(set([solicitante] + emails + superuser_emails))
                mensaje['To'] = ', '.join(destinatarios)
                mensaje['Subject'] = asunto

                bcc_destinatarios = [mi_correo]

                # Cargar la firma
                firma_path = os.path.join('media/assets/Firma', f'{user.username}.png')
                if os.path.exists(firma_path):
                    with open(firma_path, 'rb') as firma_file:
                        firma_img = MIMEImage(firma_file.read())
                        firma_img.add_header('Content-ID', '<firma>')
                        mensaje.attach(firma_img)

                # Generar el contenido HTML
                html_archivos = ""
                if urls_archivos:
                    html_archivos = f"""
                    <p>Los siguientes archivos superan el límite de 10 MB y están disponibles en los siguientes enlaces:</p>
                    <ul>
                        {''.join(f'<li><a href="{url}" target="_blank">{url}</a></li>' for url in urls_archivos)}
                    </ul>
                    """
                else:
                    html_archivos = "<p></p>"

                html_content = f"""
                <html>
                    <body>
                        <p>{formatted_message}</p>
                        <br>
                        {html_archivos}
                        <br>
                        <img src="cid:firma" alt="Firma" width="600" height="auto" />
                    </body>
                </html>
                """
                mensaje.attach(MIMEText(html_content, 'html'))

                # Adjuntar PDF
                pdf_adjunto = MIMEApplication(archivo_pdf)
                pdf_adjunto.add_header('Content-Disposition', 'attachment', filename='Ficha_de_protocolo.pdf')
                mensaje.attach(pdf_adjunto)

                # Adjuntar archivos menores a 10 MB
                for archivo_adjunto in archivos_adjuntos:
                    mensaje.attach(archivo_adjunto)

                # Configuración del servidor SMTP
                smtp_server = 'mail.munivalpo.cl'
                smtp_port = 587
                smtp_usuario = f'servervalpo\\{user.username}'
                smtp_contrasena = encotra_contraseña(user.username)

                # Enviar el correo
                server = smtplib.SMTP(smtp_server, smtp_port)
                server.starttls()
                server.login(smtp_usuario, smtp_contrasena)
                server.sendmail(
                    mi_correo,
                    destinatarios + bcc_destinatarios,  # Incluir destinatarios normales y BCC
                    mensaje.as_string()
                )
                server.quit()
                
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                        "solicitudes",
                        {"type": "send_update", "message": "actualizar"}
                    )

                return JsonResponse({'success': True})
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()  # Obtiene el error detallado
                return JsonResponse({'success': False, 'error': str(e), 'traceback': error_trace}, status=555)

                              
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@csrf_exempt
def vista_previa_reaccinacion(request, id):
    # Obtener el protocolo solicitado
    protocoloS = ProtocoloSolicitud.objects.get(id=id)
    
    # Filtrar los registros relacionados con el protocolo
    reg = Registro_designio.objects.filter(protocolo=protocoloS.id)
    
    # Construir una lista con los datos de cada registro
    registros_data = [
        {
            "id": registro.id,
            "Motivo": registro.objetivos,
            "Fecha": registro.fecha.strftime('%Y-%m-%d'),
            "Profesional_N": registro.profesional.first_name,
            "Profesional_N_1": registro.profesional.last_name,

        }
        for registro in reg
    ]
    
    # Preparar la respuesta
    data = {
        "registros": registros_data,
    }
    
    return JsonResponse(data)

def delegar_admin(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        updates = data.get('updates', [])

        for update in updates:
            user_id = update.get('user_id')
            is_superuser = update.get('is_superuser', False)

            try:
                user = User.objects.get(id=user_id)
                user.is_superuser = is_superuser
                user.save()
            
            except User.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': f'Usuario con ID {user_id} no encontrado.'}, status=404)
            
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
                "solicitudes",
                {"type": "send_update", "message": "actualizar"}
            )

        return JsonResponse({'status': 'success', 'message': 'Usuarios actualizados correctamente.'})

    usuarios = User.objects.filter(is_staff=False)
    data = {
        'Usuarios': usuarios,
    }
    return render(request, 'admin.html', data)

def encotra_contraseña(usuario, tipo='munivalpo'):
    """
    tipo: 'munivalpo' o 'gmail'
    """
    secrets_file_path = 'pass.txt'

    with open(secrets_file_path, 'r') as file:
        secrets = file.readlines()

    for secret in secrets:
        partes = secret.strip().split(':')
        if len(partes) >= 2 and partes[0] == usuario:
            if tipo == 'munivalpo':
                return partes[1]  # Contraseña municipal
            elif tipo == 'gmail' and len(partes) >= 3:
                return partes[2]  # Contraseña Gmail
    return None


@csrf_exempt
def resert_limite(request):
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            id_solicitud = data.get('id')
            solicitud = ProtocoloSolicitud.objects.get(id=id_solicitud)
            solicitud.tipo_limite = ''
            solicitud.fecha_L = None
            solicitud.save()


            return JsonResponse({'message': 'Solicitud reseteada con éxito.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Método no permitido.'}, status=405)

def solicitudes_json(request):
    # Filtrar solicitudes según si el usuario es superusuario o no
    if request.user.is_superuser:
        solicitudes = ProtocoloSolicitud.objects.all().order_by("-id")  # Orden descendente por ID
    else:
        solicitudes = ProtocoloSolicitud.objects.filter(
            Q(profesional=request.user) | Q(solicitud__profesional=request.user)
        ).distinct().order_by("-id")



    # Obtener lista de usuarios
    usuarios = list(User.objects.values('id', 'username','first_name','last_name'))

    # Lista con información procesada de solicitudes
    solicitudes_data = []

    for solicitud in solicitudes:
        # Obtener el número de designios asociados a la solicitud
        numero_designios = Registro_designio.objects.filter(protocolo=solicitud).count()

        # Obtener los apoyos asociados a la solicitud
        apoyos = Apoyo_Protocolo.objects.filter(protocolo=solicitud)
        apoyos_lista = [
            {
                "id": apoyo.profesional.id,
                "nombre": f"{apoyo.profesional.first_name} {apoyo.profesional.last_name}",
                "correo": apoyo.profesional.email,
            }
            for apoyo in apoyos
        ]

        # Calcular los días restantes hasta la fecha límite
        if solicitud.estado == "RECHAZADO":
            dias_restantes = "Rechazado"
        elif solicitud.fecha_T:
            dias_restantes = "Trabajo terminado"
        elif solicitud.fecha_L:
            # Calcular la diferencia en segundos
            total_segundos = (solicitud.fecha_L - now()).total_seconds()

            if total_segundos < 0:  # Si el tiempo ya pasó
                total_segundos = abs(total_segundos)
                dias_pasados = int(total_segundos // (24 * 3600))
                horas_pasadas = int((total_segundos % (24 * 3600)) // 3600)
                minutos_pasados = int((total_segundos % 3600) // 60)

                if dias_pasados > 0 or horas_pasadas > 0 or minutos_pasados > 0:
                    dias_restantes = f"Pasada por {dias_pasados} días, {horas_pasadas} horas y {minutos_pasados} minutos"
                else:
                    dias_restantes = "El tiempo límite ha pasado recientemente"
            else:  # Tiempo restante
                dias_habiles_restantes = calcular_dias_habiles(now().date(), solicitud.fecha_L.date())
                horas_restantes = int((total_segundos % (24 * 3600)) // 3600)
                minutos_restantes = int((total_segundos % 3600) // 60)

                if dias_habiles_restantes > 1:
                    dias_restantes = f"Te quedan {dias_habiles_restantes-1} días hábiles"
                elif horas_restantes > 0:
                    dias_restantes = f"Te quedan {horas_restantes} horas y {minutos_restantes} minutos"
                else:
                    dias_restantes = f"Te quedan {minutos_restantes} minutos"
        else:
            dias_restantes = "Sin fecha límite"

        archivos_adjuntos = ArchivoProtocolo.objects.filter(protocolo=solicitud)
        archivos_adjuntos_urls = [archivo.archivo.url for archivo in archivos_adjuntos] if archivos_adjuntos.exists() else []

        solicitudes_data.append({
            'id': solicitud.id,
            'departamento': solicitud.departamento.capitalize(),
            'direccion': solicitud.direccion.capitalize(),
            'nombre_solicitante': solicitud.nombre_solicitante,
            'nombre_proyecto': solicitud.nombre_proyecto.capitalize(),
            'corre_solicitante': solicitud.corre_solicitante,
            'area': solicitud.area.capitalize(),
            'objetivos': solicitud.objetivos.capitalize(),
            'cambios_posible': solicitud.cambios_posible.capitalize(),
            'fecha': timezone.localtime(solicitud.fecha).strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha else None,
            'codigo': solicitud.codigo,
            'orden_trabajo': solicitud.orden_trabajo,
            'fecha_D': timezone.localtime(solicitud.fecha_D).strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha_D else "Sin Fecha",
            'fecha_T': timezone.localtime(solicitud.fecha_T).strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha_T else "Sin Fecha",
            'fecha_L': solicitud.fecha_L.strftime('%Y-%m-%d') if solicitud.fecha_L else "Sin Fecha",
            'profesional_id': solicitud.profesional.id if solicitud.profesional else None,
            'profesional_nombre': f"{solicitud.profesional.first_name} {solicitud.profesional.last_name}" if solicitud.profesional else "Sin asignar",
            'profesional_correo': solicitud.profesional.email if solicitud.profesional else "Sin asignar",
            'order_trabajo': solicitud.orden_trabajo if solicitud.orden_trabajo else "Sin asignar",
            'tipo_limite': solicitud.tipo_limite,
            'estado': solicitud.estado,
            'enviado_correo': solicitud.enviado_correo,
            'enviado_correo_t': solicitud.enviado_correo_t,
            'numero_designios': numero_designios,
            'dias_restantes': dias_restantes,
            'archivos_adjuntos_urls': archivos_adjuntos_urls,
            'apoyos': apoyos_lista,  # Agregar la lista de apoyos
        })



    # Estructurar la respuesta en JSON
    data = {
        'id_user': request.user.id,  # Indica si el usuario es superusuario
        'es_superuser': request.user.is_superuser,  # Indica si el usuario es superusuario

        'OPCIONES': OPCIONES,  # Opciones de selección
        'Solicitudes': solicitudes_data,  # Datos de solicitudes con información adicional
        'Usuarios': usuarios  # Lista de usuarios
    }

    

    return JsonResponse(data, safe=False)

def actualizar_limite_solicitud(request):
    if request.method == 'POST':
        if not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'No tienes permisos para cambiar la carga de trabajo'}, status=403)

        solicitud_id = request.POST.get('id')
        nuevo_limite = request.POST.get('tipo_limite')

        try:
            solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)
            solicitud.tipo_limite = nuevo_limite
            solicitud.save()

            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                    "solicitudes",
                    {"type": "send_update", "message": "actualizar"}
                )
            return JsonResponse({'success': True, 'message': 'Carga de trabajo actualizada correctamente'})
        except ProtocoloSolicitud.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Solicitud no encontrada'})

    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

def detalle_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(ProtocoloSolicitud, id=solicitud_id)

    data = {
        'id': solicitud.id,
        'nombre_solicitante': solicitud.nombre_solicitante,
        'fecha': solicitud.fecha.strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha else 'No disponible',
        'departamento': solicitud.departamento,
        'estado': solicitud.estado,
        'numero_designios': solicitud.registro_designio_set.count(),  # Contar los registros relacionados
        'dias_restantes': "Sin fecha límite" if not solicitud.fecha_T else "Trabajo terminado" if solicitud.fecha_T else "En proceso"
    }

    return JsonResponse(data)

def solicitud_express(request):
    if request.method == "POST":
        try:
            with transaction.atomic():
                Protocolo = ProtocoloSolicitud.objects.create(
                    direccion=request.POST['direccion'],
                    departamento=request.POST['departamento'],
                    nombre_solicitante=request.POST['nombre_solicitante'].title(),
                    nombre_proyecto=request.POST['nombre_proyecto'],
                    corre_solicitante=request.POST['corre_solicitante'],
                    area=request.POST['area'],
                    objetivos=request.POST['objetivos'],
                    cambios_posible=request.POST['cambios_posible'],
                    anexo=request.POST['anexo'],
                )
                archivos_adjuntos = request.FILES.getlist('archivo')
                if archivos_adjuntos:
                    for archivo in archivos_adjuntos:
                        ArchivoProtocolo.objects.create(protocolo=Protocolo, archivo=archivo)


                insumo_ids = [int(insumo.split('_')[1]) for insumo in request.POST.getlist('insumo') if insumo.startswith('insumo_')]
                
                if not insumo_ids:
                    print("No se seleccionaron insumos válidos.")
                    return HttpResponse("No se seleccionaron insumos válidos.", status=400)

                insumos_a_guardar = []
                for insumo_id in insumo_ids:
                    try:
                        insumo = Insumo.objects.get(id=insumo_id)
                        print(f"Insumo encontrado con ID {insumo_id}: {insumo}")
                    except Insumo.DoesNotExist:
                        print(f"Insumo con ID {insumo_id} no encontrado, creando nuevo insumo.")
                        insumo = Insumo.objects.create(id=insumo_id, nombre=f"Insumo {insumo_id}")
                    insumos_a_guardar.append(insumo)

                Protocolo.insumo.set(insumos_a_guardar)

                producto_ids = [int(producto.split('_')[1]) for producto in request.POST.getlist('producto') if producto.startswith('producto_')]
                print("Producto IDs extraídos:", producto_ids)
                
                Protocolo.codigo = str(Protocolo.id)
                Protocolo.save()

                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    "solicitudes",
                    {"type": "send_update", "message": "actualizar"}
                )


            archivos_adjuntos = request.FILES.getlist('archivo')
            for archivo in archivos_adjuntos:
                ArchivoProtocolo.objects.create(protocolo=Protocolo, archivo=archivo)

            
            return JsonResponse({'success': True, 'message': 'Se creó la solicitud correctamente'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error al crear una solicitud: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

def usuarios_disponibles(request):
    protocolo_id = request.GET.get('protocolo_id')

    try:
        protocolo = ProtocoloSolicitud.objects.get(id=protocolo_id)
        # Excluir el profesional asignado principal
        profesional_id = protocolo.profesional.id if protocolo.profesional else None

        # Obtener todos los usuarios, excluyendo el profesional principal
        usuarios = User.objects.exclude(id=profesional_id,) .values('id', 'first_name', 'last_name')

        # Obtener los apoyos ya registrados para este protocolo
        apoyo_qs = Apoyo_Protocolo.objects.filter(protocolo=protocolo)
        apoyo_user_ids = list(apoyo_qs.values_list('profesional_id', flat=True))

        usuarios_list = []
        for u in usuarios:
            u['ya_agregado'] = u['id'] in apoyo_user_ids
            usuarios_list.append(u)

        return JsonResponse({"usuarios": usuarios_list}, safe=False)

    except ProtocoloSolicitud.DoesNotExist:
        return JsonResponse({"error": "Protocolo no encontrado"}, status=404)

    
@csrf_exempt
def agregar_apoyo(request):
    if request.method == "POST":
        protocolo_id = request.POST.get("protocolo_id")
        usuarios_json = request.POST.get("usuarios", "[]")
        usuarios_ids = json.loads(usuarios_json)

        try:
            protocolo = ProtocoloSolicitud.objects.get(id=protocolo_id)

            # Obtiene todos los apoyos actuales para este protocolo
            apoyos_actuales = Apoyo_Protocolo.objects.filter(protocolo=protocolo)
            # Extrae los IDs de los usuarios que ya están asignados como apoyo
            apoyos_ids_actuales = list(apoyos_actuales.values_list('profesional_id', flat=True))

            # Elimina aquellos apoyos que ya no están en la lista seleccionada
            Apoyo_Protocolo.objects.filter(protocolo=protocolo).exclude(profesional_id__in=usuarios_ids).delete()

            # Agrega los nuevos apoyos que no existan ya
            for usuario_id in usuarios_ids:
                if usuario_id not in apoyos_ids_actuales:
                    usuario = User.objects.get(id=usuario_id)
                    Apoyo_Protocolo.objects.create(protocolo=protocolo, profesional=usuario)
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "solicitudes",
                {"type": "send_update", "message": "actualizar"}
            )

            return JsonResponse({"success": True})

        except ProtocoloSolicitud.DoesNotExist:
            return JsonResponse({"error": "Protocolo no encontrado"}, status=404)
        except User.DoesNotExist:
            return JsonResponse({"error": "Algún usuario no se encontró"}, status=404)

    return JsonResponse({"error": "Método no permitido"}, status=405)


def apoyo_trabajo(request):
    protocolo_id = request.GET.get("protocolo_id")

    if not protocolo_id:
        return JsonResponse({"error": "Protocolo no encontrado"}, status=400)

    try:
        protocolo = ProtocoloSolicitud.objects.get(id=protocolo_id)
        apoyos = Apoyo_Protocolo.objects.filter(protocolo=protocolo)

        apoyo_data = [
            {
                "id": apoyo.profesional.id,
                "first_name": apoyo.profesional.first_name,
                "last_name": apoyo.profesional.last_name,
                "ya_agregado": True
            }
            for apoyo in apoyos
        ]

        return JsonResponse({"apoyos": apoyo_data})

    except ProtocoloSolicitud.DoesNotExist:
        return JsonResponse({"error": "Protocolo no encontrado"}, status=404)

@csrf_exempt  # Deshabilita CSRF solo para pruebas (en producción usa @csrf_protect y pasa el token en AJAX)


@csrf_exempt
def obtener_nota(request):
    protocolo_id = request.GET.get("protocolo_id")

    if not protocolo_id:
        return JsonResponse({"error": "Protocolo no encontrado"}, status=400)

    try:
        protocolo = ProtocoloSolicitud.objects.get(id=protocolo_id)
        apoyos = Apoyo_Protocolo.objects.filter(protocolo=protocolo)

        # Obtener valores en formato 1-100 (convertir de 0.XX a XX)
        porcentaje_total = (protocolo.valor_de_trabajo or 0) * 100
        porcentaje_profesional = (protocolo.valor_de_trabajo_funcionario or 0) * 100
        porcentaje_apoyo = sum(
            apoyo.valor_de_trabajo if apoyo.valor_de_trabajo is not None else 0
            for apoyo in apoyos
        ) * 100

        # Construir respuesta JSON
        apoyo_data = [
            {
                "id": apoyo.profesional.id,
                "first_name": apoyo.profesional.first_name,
                "last_name": apoyo.profesional.last_name,
                "ya_agregado": True
            }
            for apoyo in apoyos
        ]

        return JsonResponse({
            "porcentaje_total": int(porcentaje_total),  # Convertir a entero
            "porcentaje_profesional": int(porcentaje_profesional),  # Convertir a entero
            "porcentaje_apoyo": int(porcentaje_apoyo),  # Convertir a entero
            "apoyos": apoyo_data
        })

    except ProtocoloSolicitud.DoesNotExist:
        return JsonResponse({"error": "Protocolo no encontrado"}, status=404)


from django.shortcuts import render, redirect
from django.utils.timezone import now, is_naive, make_aware
from datetime import datetime
from .models import Solicitud
from django.db.models.functions import ExtractYear
from django.utils.dateparse import parse_datetime
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

def capitalize_words(texto):
    return ' '.join(p.capitalize() for p in texto.split())

def vista_consolidada(request):
    mensaje_exito = ""
    mensaje_error = ""

    # FILTROS GET
    solicitudes = Solicitud.objects.all()

    id_filtro = request.GET.get('id')
    if id_filtro:
        solicitudes = solicitudes.filter(id__icontains=id_filtro)

    fecha_subida_inicio = request.GET.get('fecha_subida_inicio')
    fecha_subida_fin = request.GET.get('fecha_subida_fin')
    if fecha_subida_inicio:
        solicitudes = solicitudes.filter(fecha_subida__date__gte=fecha_subida_inicio)
    if fecha_subida_fin:
        solicitudes = solicitudes.filter(fecha_subida__date__lte=fecha_subida_fin)

    fecha_actualizacion_inicio = request.GET.get('fecha_actualizacion_inicio')
    fecha_actualizacion_fin = request.GET.get('fecha_actualizacion_fin')
    if fecha_actualizacion_inicio:
        solicitudes = solicitudes.filter(fecha_actualizacion__date__gte=fecha_actualizacion_inicio)
    if fecha_actualizacion_fin:
        solicitudes = solicitudes.filter(fecha_actualizacion__date__lte=fecha_actualizacion_fin)

    # PROCESO POST (igual como tenías antes)
    if request.method == "POST" and "solicitud_id" in request.POST:
        solicitud_id = request.POST.get("solicitud_id")
        try:
            solicitud = Solicitud.objects.get(id=solicitud_id)
        except Solicitud.DoesNotExist:
            mensaje_error = "Solicitud no encontrada."
        else:
            comentario = capitalize_words(request.POST.get("comentario", "").strip())
            fecha_actualizacion_str = request.POST.get("fecha_actualizacion", "").strip()
            validado_por = capitalize_words(request.POST.get("validado_por", "").strip())

            if not comentario or not fecha_actualizacion_str or not validado_por:
                mensaje_error = "Debe completar todos los campos para actualizar."
            else:
                try:
                    fecha_actualizacion = datetime.strptime(fecha_actualizacion_str, '%Y-%m-%dT%H:%M')
                    if is_naive(fecha_actualizacion):
                        fecha_actualizacion = make_aware(fecha_actualizacion)
                except (ValueError, TypeError):
                    mensaje_error = "Formato de fecha no válido."
                else:
                    solicitud.comentario = comentario
                    solicitud.fecha_actualizacion = fecha_actualizacion
                    solicitud.validado_por = validado_por
                    solicitud.save()
                    return redirect("vista_consolidada")

    elif request.method == "POST":
        nombre_usuario = capitalize_words(request.POST.get("nombre_usuario", "").strip())
        nombre_archivo = capitalize_words(request.POST.get("nombre_archivo", "").strip())
        fuente = capitalize_words(request.POST.get("fuente", "").strip())
        lugar = capitalize_words(request.POST.get("lugar", "").strip())
        fecha_actualizacion_str = request.POST.get("fecha_actualizacion", "").strip()
        validado_por = capitalize_words(request.POST.get("validado_por", "Ninguno").strip())
        estado = request.POST.get("estado", 1)

        if not (nombre_usuario and nombre_archivo and fuente and lugar and fecha_actualizacion_str):
            mensaje_error = "Debe completar todos los campos para crear una solicitud."
        else:
            try:
                fecha_actualizacion = datetime.strptime(fecha_actualizacion_str, '%Y-%m-%dT%H:%M')
                if is_naive(fecha_actualizacion):
                    fecha_actualizacion = make_aware(fecha_actualizacion)
            except (ValueError, TypeError):
                mensaje_error = "Formato de fecha no válido."
            else:
                Solicitud.objects.create(
                    nombre_usuario=nombre_usuario,
                    nombre_archivo=nombre_archivo,
                    fuente=fuente,
                    lugar=lugar,
                    fecha_actualizacion=fecha_actualizacion,
                    validado_por=validado_por,
                    estado=estado,
                    fecha_subida=now(),
                )
                return redirect("vista_consolidada")

    # Procesar días restantes (exactamente igual)
    for solicitud in solicitudes:
        delta = solicitud.fecha_actualizacion - now()
        dias_restantes = delta.days
        solicitud.dias_restantes = dias_restantes
        solicitud.dias_restantes_abs = abs(dias_restantes)

    anios_qs = Solicitud.objects.annotate(anio=ExtractYear('fecha_actualizacion')).values_list('anio', flat=True).distinct().order_by('-anio')
    anios = list(anios_qs)

    context = {
        "solicitudes": solicitudes,
        "mensaje_exito": mensaje_exito,
        "mensaje_error": mensaje_error,
        "today": now().strftime("%Y-%m-%dT%H:%M"),
        "anios": anios,
    }
    return render(request, "consolidado/consolidado.html", context)

@csrf_exempt
def actualizar_estado(request, id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            nuevo_estado = data.get("estado")
            nueva_fecha_str = data.get("fecha_actualizacion")
            validado_por = data.get("validado_por", "").strip()
            comentario = data.get("comentario", "").strip()

            solicitud = Solicitud.objects.get(id=id)
            nueva_fecha = parse_datetime(nueva_fecha_str)
            if not nueva_fecha:
                return JsonResponse({"error": "Fecha de actualización inválida."}, status=400)

            validado_por = capitalize_words(validado_por)
            comentario = capitalize_words(comentario)

            solicitud.estado = nuevo_estado
            solicitud.fecha_actualizacion = nueva_fecha
            solicitud.validado_por = validado_por
            solicitud.comentario = comentario
            solicitud.save()

            return JsonResponse({"mensaje": "Estado actualizado correctamente."})
        except Solicitud.DoesNotExist:
            return JsonResponse({"error": "Solicitud no encontrada."}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Método no permitido"}, status=405)
